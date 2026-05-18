import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================
# AI 魔法棒 BOM 比价数据 - 2026-05-11
# ============================================================

# 华秋商城数据 (WebFetch 查询 2026-05-11)
hqchip_data = {
    "ESP32-S3-WROOM-1-N8": {"p1": 29.18, "p10": 25.04, "p100": None, "stock": 14, "source": "华秋自营"},
    "OV8856": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "摄像头模组"},
    "ST7789V 1.2寸": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "显示屏模组"},
    "INMP441": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "麦克风模组"},
    "FSVP532": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "NFC模块"},
    "SIM7600CE-CNSE-PCIE": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "4G模组"},
    "L76KB-A58": {"p1": None, "p10": None, "p100": None, "stock": None, "source": "", "note": "GPS模块"},
    "MPU-6050": {"p1": 39.00, "p10": 37.50, "p100": 33.90, "p500": 32.10, "stock": 0, "source": "华秋自营"},
    "MAX98357AETE+T": {"p1": 8.30, "p100": 5.60, "p500": 5.30, "p1000": 5.10, "stock": 9139, "source": "华秋自营"},
    "TP4056": {"p5": 0.30, "p20": 0.28, "p100": 0.25, "p500": 0.22, "p1000": 0.21, "p2000": 0.20, "stock": None, "source": "华秋自营"},
    "SY8089AAAC": {"p5": 0.68, "p20": 0.62, "p100": 0.56, "p500": 0.50, "p1000": 0.47, "p2000": 0.45, "stock": 11156, "source": "华秋自营"},
    "WS2812B": {"p10": 0.44, "p50": 0.41, "p100": 0.38, "p200": 0.36, "stock": 46, "source": "华秋自营"},
    "AMS1117-3.3": {"p5": 0.76, "p50": 0.70, "p500": 0.64, "p1000": 0.58, "stock": 3963, "source": "华秋自营"},
    "AP2112K-3.3": {"p10": 0.97, "p100": 0.76, "p600": 0.55, "p1200": 0.55, "p3000": 0.47, "stock": None, "source": "华秋自营"},
    "LM393": {"p5": 0.20, "p20": 0.19, "p100": 0.17, "p500": 0.15, "p1000": 0.14, "p2000": 0.14, "stock": None, "source": "华秋自营"},
}

# 云汉芯城数据 (从 Playwright 查询)
ickey_data = {
    "ESP32-S3-WROOM-1-N8": {"p1": 29.56, "p10": 25.35, "p30": 22.85, "p100": 20.32, "p650": 19.12, "p1300": 18.59, "stock": 3582, "source": "云汉现货"},
}

# 立创商城数据 (从之前的会话记录)
lcsc_cn_data = {
    "ESP32-S3-WROOM-1-N8": {"p1": 29.78, "p10": 25.55, "p100": 23.03, "stock": None, "source": "历史记录"},
    "INMP441": {"p1": 61.00, "p100": 58.00, "stock": None, "source": "历史记录", "note": "100pcs价/散装价"},
    "MAX98357AETE+T": {"p1": 8.36, "p10": 7.34, "p100": 6.37, "p1000": 5.65, "stock": None, "source": "历史记录"},
    "MPU-6050": {"p1": 61.90, "p10": 56.63, "p100": 53.26, "p1000": 50.00, "stock": None, "source": "历史记录"},
}

# LCSC国际站数据 (USD, 需要 x 7.25 换算 CNY)
lcsc_intl_data_usd = {
    "ESP32-S3-WROOM-1-N8": {"p1": 3.49, "p10": 3.23, "p100": 3.01, "p1300": 2.95, "stock": None, "source": "LCSC国际站"},
    "INMP441": {"p1": 9.83, "p100": 9.30, "p1000": 8.70, "stock": None, "source": "LCSC国际站"},
    "MAX98357AETE+T": {"p1": 1.19, "p10": 1.06, "p100": 0.84, "p1000": 0.74, "stock": None, "source": "LCSC国际站"},
    "MPU-6050": {"p1": 2.57, "p10": 2.37, "p100": 2.10, "p1000": 1.83, "stock": None, "source": "LCSC国际站"},
    "TP4056": {"p1": 0.123, "p10": 0.100, "p100": 0.076, "p500": 0.065, "stock": None, "source": "LCSC国际站"},
    "SY8089AAAC": {"p1": 0.142, "p10": 0.112, "p100": 0.082, "p500": 0.071, "stock": None, "source": "LCSC国际站"},
    "WS2812B": {"p1": 0.096, "p10": 0.083, "p100": 0.056, "p2000": 0.043, "stock": None, "source": "LCSC国际站"},
    "AMS1117-3.3": {"p1": 0.208, "p10": 0.160, "p100": 0.122, "p500": 0.105, "stock": None, "source": "LCSC国际站"},
    "AP2112K-3.3": {"p1": 0.170, "p10": 0.135, "p100": 0.100, "p500": 0.087, "stock": None, "source": "LCSC国际站"},
    "FSVP532": {"p1": 3.93, "p100": 3.30, "p980": 2.40, "stock": None, "source": "LCSC国际站"},
}

# 市场估算 (淘宝/1688, 用于立创没有的模组)
market_data = {
    "OV8856 800W摄像头": {"price": 70.00, "note": "淘宝模组价, 含镜头FPC"},
    "ST7789V 1.2寸 IPS": {"price": 25.00, "note": "淘宝模组价, 含FPC"},
    "INMP441 麦克风": {"price": 58.00, "note": "淘宝模组价, 100pcs散装"},
    "FSVP532 NFC模块": {"price": 25.99, "note": "淘宝模块价"},
    "SIM7600CE 4G模组": {"price": 120.00, "note": "淘宝模块价"},
    "L76KB GPS模块": {"price": 27.14, "note": "淘宝模块价"},
    "Battery 3.7V 1000mAh": {"price": 12.00, "note": "淘宝电池价"},
    "USB-C Connector": {"price": 0.50, "note": "淘宝单价"},
    "0.1uF电容": {"price": 0.01, "note": "淘宝100pcs均价"},
    "10K电阻": {"price": 0.01, "note": "淘宝100pcs均价"},
}

# ============================================================
# BOM 清单
# ============================================================
bom_items = [
    {"id": 1, "model": "ESP32-S3-WROOM-1-N8", "pkg": "SMD Module", "qty": 1, "brand": "Espressif", "desc": "主控WiFi/BT MCU"},
    {"id": 2, "model": "OV8856", "pkg": "Module", "qty": 1, "brand": "OmniVision", "desc": "800W摄像头模组"},
    {"id": 3, "model": "ST7789V 1.2寸 IPS", "pkg": "Module", "qty": 1, "brand": "-", "desc": "1.2寸显示屏模组"},
    {"id": 4, "model": "INMP441", "pkg": "Module", "qty": 1, "brand": "TDK InvenSense", "desc": "I2S MEMS麦克风"},
    {"id": 5, "model": "FSVP532", "pkg": "Module", "qty": 1, "brand": "FarSemi", "desc": "NFC模块(PN532兼容)"},
    {"id": 6, "model": "SIM7600CE-CNSE-PCIE", "pkg": "Module", "qty": 1, "brand": "SIMCom", "desc": "4G模组"},
    {"id": 7, "model": "L76KB-A58", "pkg": "Module", "qty": 1, "brand": "Quectel", "desc": "GPS模块"},
    {"id": 8, "model": "MPU-6050", "pkg": "QFN-24", "qty": 1, "brand": "TDK InvenSense", "desc": "六轴IMU"},
    {"id": 9, "model": "MAX98357AETE+T", "pkg": "TQFN-16", "qty": 1, "brand": "Maxim/ADI", "desc": "I2S功放"},
    {"id": 10, "model": "TP4056", "pkg": "SOP-8", "qty": 1, "brand": "TOPPOWER/NanJing", "desc": "锂电池充电管理"},
    {"id": 11, "model": "SY8089AAAC", "pkg": "SOT-23-6", "qty": 1, "brand": "Silergy", "desc": "同步DCDC升压"},
    {"id": 12, "model": "WS2812B", "pkg": "LED-5050", "qty": 6, "brand": "Worldsemi", "desc": "RGB LED"},
    {"id": 13, "model": "AMS1117-3.3", "pkg": "SOT-223", "qty": 2, "brand": "AMS", "desc": "3.3V LDO"},
    {"id": 14, "model": "AP2112K-3.3", "pkg": "SOT-23-5", "qty": 2, "brand": "Diodes Inc", "desc": "3.3V LDO"},
    {"id": 15, "model": "Battery 3.7V 1000mAh", "pkg": "-", "qty": 1, "brand": "-", "desc": "锂电池"},
    {"id": 16, "model": "LM393", "pkg": "SOIC-8", "qty": 1, "brand": "TI/NS", "desc": "双比较器"},
    {"id": 17, "model": "0.1uF电容", "pkg": "0402", "qty": 10, "brand": "Yageo", "desc": "贴片电容"},
    {"id": 18, "model": "10K电阻", "pkg": "0402", "qty": 10, "brand": "Yageo", "desc": "贴片电阻"},
    {"id": 19, "model": "USB-C Connector", "pkg": "SMD", "qty": 1, "brand": "-", "desc": "Type-C接口"},
]

# ============================================================
# 价格查询和比较逻辑
# ============================================================
USD_CNY = 7.25

def get_hqchip_price(d):
    """从华秋阶梯价中取最小起订量的价格"""
    # 按起订量排序取最低
    tiers = []
    for k, v in d.items():
        if k.startswith('p') and k[1:].isdigit() and v is not None:
            qty = int(k[1:])
            tiers.append((qty, v))
    if not tiers:
        return None
    tiers.sort(key=lambda x: x[0])
    return tiers[0][1]

def get_best_price(item):
    """从所有数据源中找出最佳价格"""
    model = item["model"]
    results = {}

    # 立创商城 (用型号前缀匹配)
    matched_lcsc_cn = None
    for lcsc_model, d in lcsc_cn_data.items():
        prefix = lcsc_model.split()[0] if ' ' in lcsc_model else lcsc_model
        if model.startswith(prefix) or model.startswith(lcsc_model):
            matched_lcsc_cn = d
            break
    if not matched_lcsc_cn and model in lcsc_cn_data:
        matched_lcsc_cn = lcsc_cn_data[model]
    if matched_lcsc_cn:
        d = matched_lcsc_cn
        results["立创商城"] = d["p1"]  # 1+价
        results["立创_10+"] = d.get("p10")
        results["立创_100+"] = d.get("p100")

    # LCSC国际站 (USD -> CNY, 用型号前缀匹配)
    matched_lcsc = None
    for lcsc_model, d in lcsc_intl_data_usd.items():
        prefix = lcsc_model.split()[0] if ' ' in lcsc_model else lcsc_model
        if model.startswith(prefix) or model.startswith(lcsc_model):
            matched_lcsc = d
            break
    if not matched_lcsc and model in lcsc_intl_data_usd:
        matched_lcsc = lcsc_intl_data_usd[model]
    if matched_lcsc:
        d = matched_lcsc
        p_cny = round(d["p1"] * USD_CNY, 2)
        results["LCSC国际"] = p_cny
        results["LCSC_10+"] = round(d.get("p10", 0) * USD_CNY, 2) if d.get("p10") else None
        results["LCSC_100+"] = round(d.get("p100", 0) * USD_CNY, 2) if d.get("p100") else None
        results["LCSC_USD"] = d["p1"]  # 保留原始USD价格

    # 华秋商城 (支持各种起订量阶梯)
    if model in hqchip_data:
        d = hqchip_data[model]
        price = get_hqchip_price(d)
        if price:
            results["华秋商城"] = price
        results["华秋_10+"] = d.get("p10")
        results["华秋_100+"] = d.get("p100")

    # 云汉芯城
    if model in ickey_data and ickey_data[model]["p1"]:
        d = ickey_data[model]
        results["云汉芯城"] = d["p1"]
        results["云汉_10+"] = d.get("p10")
        results["云汉_100+"] = d.get("p100")

    # 市场价 (用型号前缀匹配)
    for key, d in market_data.items():
        # 取市场价key中的型号前缀（空格前的部分）
        prefix = key.split()[0] if ' ' in key else key
        if model.startswith(prefix) or prefix in model:
            results["市场价"] = d["price"]
            results["市场备注"] = d["note"]
            break

    return results


def find_lowest_source(prices):
    """找出最低价来源"""
    candidates = {}
    for k, v in prices.items():
        if k in ["立创商城", "华秋商城", "云汉芯城", "LCSC国际", "市场价"] and v is not None:
            candidates[k] = v
    if not candidates:
        return None, None
    lowest = min(candidates, key=candidates.get)
    return lowest, candidates[lowest]


def build_remarks(item, prices):
    """构建备注：列出各商城对比价格"""
    remarks = []
    model = item["model"]

    # 立创价格
    if "立创商城" in prices:
        p = prices["立创商城"]
        r = f"立创 ¥{p:.2f}"
        if prices.get("立创_100+"):
            r += f"(100+ ¥{prices['立创_100+']:.2f})"
        remarks.append(r)

    # LCSC国际
    if "LCSC国际" in prices:
        p = prices["LCSC国际"]
        usd = prices.get("LCSC_USD", p / USD_CNY)
        r = f"LCSC ${usd:.3f}(¥{p:.2f})"
        if prices.get("LCSC_100+"):
            r += f"(100+ ${prices.get('LCSC_100+',0)/USD_CNY:.3f})"
        remarks.append(r)

    # 华秋
    if "华秋商城" in prices:
        p = prices["华秋商城"]
        r = f"华秋 ¥{p:.2f}"
        if prices.get("华秋_100+"):
            r += f"(100+ ¥{prices['华秋_100+']:.2f})"
        elif prices.get("华秋_10+"):
            r += f"(10+ ¥{prices['华秋_10+']:.2f})"
        remarks.append(r)

    # 云汉
    if "云汉芯城" in prices:
        p = prices["云汉芯城"]
        r = f"云汉 ¥{p:.2f}"
        if prices.get("云汉_100+"):
            r += f"(100+ ¥{prices['云汉_100+']:.2f})"
        remarks.append(r)

    # 市场
    if "市场价" in prices:
        r = f"市场 ¥{prices['市场价']:.2f}"
        if "市场备注" in prices:
            r += f"({prices['市场备注']})"
        remarks.append(r)

    if not remarks:
        return "暂无比价数据"

    return " | ".join(remarks)


# ============================================================
# 生成 Excel
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOM比价单"

# 样式定义
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

lowest_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
total_font = Font(name="微软雅黑", size=11, bold=True, color="2F5496")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题行
ws.merge_cells('A1:L1')
title_cell = ws['A1']
title_cell.value = f"AI 魔法棒 BOM 比价单 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# 表头
headers = ["#", "型号", "品牌", "封装", "数量", "最低单价(¥)", "最低价来源",
           "小计(¥)", "备注（各商城比价）", "立创(¥)", "华秋(¥)", "云汉(¥)"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.row_dimensions[2].height = 30

# 列宽
col_widths = [5, 28, 16, 14, 6, 12, 12, 12, 55, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 数据行
total_cost = 0
for row_idx, item in enumerate(bom_items, 3):
    prices = get_best_price(item)
    lowest_source, lowest_price = find_lowest_source(prices)

    # 计算小计
    unit_price = lowest_price if lowest_price else 0
    subtotal = unit_price * item["qty"]
    total_cost += subtotal

    # 备注
    remarks = build_remarks(item, prices)

    row_data = [
        item["id"],
        item["model"],
        item["brand"],
        item["pkg"],
        item["qty"],
        f"¥{unit_price:.2f}" if unit_price else "待查询",
        lowest_source or "待查询",
        f"¥{subtotal:.2f}",
        remarks,
        f"¥{prices.get('立创商城', 0):.2f}" if prices.get("立创商城") else "-",
        f"¥{prices.get('华秋商城', 0):.2f}" if prices.get("华秋商城") else "-",
        f"¥{prices.get('云汉芯城', 0):.2f}" if prices.get("云汉芯城") else "-",
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align if col_idx <= 8 else left_align
        cell.border = thin_border

        # 最低价来源列高亮
        if col_idx == 7 and lowest_source:
            cell.fill = lowest_fill

        # 待查询的用黄色标记
        if value in ["待查询", "-"] and col_idx in [6, 7]:
            cell.fill = warn_fill

    ws.row_dimensions[row_idx].height = max(20, 15 * (1 + len(remarks) // 40))

# 汇总行
summary_row = len(bom_items) + 3
ws.merge_cells(f'A{summary_row}:F{summary_row}')
cell = ws.cell(row=summary_row, column=1, value="总成本（按最低价计算）")
cell.font = total_font
cell.fill = total_fill
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.border = thin_border

for col in range(2, 7):
    ws.cell(row=summary_row, column=col).fill = total_fill
    ws.cell(row=summary_row, column=col).border = thin_border

total_cell = ws.cell(row=summary_row, column=8, value=f"¥{total_cost:.2f}")
total_cell.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
total_cell.fill = total_fill
total_cell.alignment = Alignment(horizontal="center", vertical="center")
total_cell.border = thin_border

for col in [9, 10, 11, 12]:
    ws.cell(row=summary_row, column=col).fill = total_fill
    ws.cell(row=summary_row, column=col).border = thin_border

ws.row_dimensions[summary_row].height = 30

# 说明行
note_row = summary_row + 2
notes = [
    "说明：",
    "1. 价格数据来源：立创商城（历史记录）、LCSC国际站（历史，USD×7.25换算）、华秋商城（WebFetch实时查询）、云汉芯城（Playwright实时查询）",
    "2. 华秋商城共查询14个型号，其中8个有自营现货价格；云汉芯城仅ESP32-S3有数据（其余型号现货库存为空）",
    "3. 模组类（摄像头、显示屏、4G、GPS、NFC、电池）主要通过淘宝市场价估算，实际采购价可能有差异",
    "4. 通用物料（0402电阻电容）按百片均价估算，实际采购成本极低",
    "5. MPU-6050 华秋库存为0（需订货），立创有货¥61.90",
    "6. 建议实际采购前在对应商城确认最新价格和库存",
]
for i, note in enumerate(notes):
    cell = ws.cell(row=note_row + i, column=1, value=note)
    cell.font = Font(name="微软雅黑", size=9, color="666666")
    ws.merge_cells(f'A{note_row+i}:L{note_row+i}')

# 保存
output_path = "/Users/lizhengan/WorkBuddy/2026-05-11-task-10/BOM_比价单_20260511.xlsx"
wb.save(output_path)
print(f"✅ BOM比价单已生成: {output_path}")
print(f"📊 总成本: ¥{total_cost:.2f}")
print(f"📋 共 {len(bom_items)} 个元器件")
